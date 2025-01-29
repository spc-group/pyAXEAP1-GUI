# :author: Alexander Berno

"""File Loading classes.

Includes base LoadFile class and all subclasses.

Each class has a fileDialog function and a loadData function.
fileDialog uses Qt dialog windows to get file names.
loadData loads and returns data in a readible form."""

from PyQt6.QtWidgets import QFileDialog
import pathlib
from calibFunctions import loadCalib, getCoordsFromScans
from openpyxl import load_workbook
from ErrorWindow import ErrorWindow
import axeap.core as core

from tiled.client import from_uri
from tqdm.notebook import tqdm
import h5py
from rich import print as rprint
from rich.tree import Tree
from collections.abc import Sequence

desktop_directory = str(pathlib.Path.home() / "Desktop")


class LoadFile:
    """Base file loading class."""

    def fileDialog(parent: any): ...

    def loadData(directory: str | pathlib.Path): ...


class LoadTiffCalib(LoadFile):
    def fileDialog(parent: any):
        direct = QFileDialog.getOpenFileNames(
            parent=parent,
            directory=desktop_directory,
            filter="TIF Files (*.tif *.tiff)",
        )
        if direct[0] != [] and direct[1] != "":
            return direct[0]
        else:
            return None

    def loadData(directory: str):
        return loadCalib(directory)


class LoadH5Data(LoadFile):
    def fileDialog(parent: any):
        direct = QFileDialog.getOpenFileNames(
            parent=parent,
            directory=desktop_directory,
            filter="NX Files (*.nx *.*)",
        )
        if direct[0] != [] and direct[1] != "":
            return direct[0]
        else:
            return None

    def loadData(directory: str | tuple):
        # "C:\\Users\\bernoa\\Desktop\\mark_data\\example_count_data.nx"

        def getImages(node):
            images = []
            for key in node.keys():
                child_node = node[key]
                if hasattr(child_node, "dtype"):
                    if key == "eiger_image":
                        images.append(child_node)
                    elif key.endswith("_image"):
                        images.append(child_node)
                elif hasattr(child_node, "keys"):
                    im = getImages(child_node)
                    for i in im:
                        images.append(i)
            return images

        points = []
        if isinstance(directory, Sequence) and not isinstance(directory, (str,)):
            for d in directory:
                points += LoadH5Data.loadData(d)

        else:
            with h5py.File(directory, mode="r") as fd:
                images = getImages(fd)
                for imgs in images:
                    for img in imgs:
                        points.append(img[0])
        return points


class LoadInfoData(LoadFile):
    def fileDialog(parent: any):
        info_file = QFileDialog.getOpenFileName(
            parent=parent, directory=desktop_directory
        )
        return info_file

    def loadData(parent: any, directory: str, rtype="calib"):
        old = parent.info_file
        if rtype == "calib":
            scans = parent.calibscans
            energies = parent.calib_energies
        if directory[0][-5:] == ".xlsx":
            wb = load_workbook(directory[0], read_only=True)
            ws = wb.worksheets[0]
            if rtype == "calib":
                for i, scan in enumerate(scans.items):
                    line = ws.cell(i + 1, 1).value
                    scan.meta["IncidentEnergy"] = line
                    energies[i].changeVal(line)
            elif rtype == "rxes":
                table = []
                lines = []
                empty = False
                i = 1
                j = 1
                while not empty:
                    line = ws.cell(i, j).value
                    if len(line):
                        lines.append(line)
                    else:
                        if i == 1:
                            empty = True
                        else:
                            table.append(lines.copy())
                            lines = []
                            j += 1
                            i = 0
                    i += 1
            wb.close()

        elif directory[0][-4:] == ".txt":
            with open(directory[0], "r") as f:
                lines = f.readlines()
                for i, line in enumerate(lines):
                    # scan.meta["IncidentEnergy"] = lines[i]
                    energies[i].changeVal(float(line[: line.find("\n")]))

        elif directory[1] != "":
            if rtype == "calib":
                try:
                    scans.addCalibRunInfo(core.CalibRunInfo(directory[0]))
                    for i, e in enumerate(energies):
                        e.changeVal(scans.items[i].meta["IncidentEnergy"])

                except Exception or Warning:
                    directory = old
                    parent.error = ErrorWindow("badInfoFile")
            elif rtype == "rxes":
                info = core.CalibRunInfo(directory[0])
                table = info._table

        else:
            directory = old

        if rtype == "rxes":
            return directory, table

        return directory


class LoadTifSpectraData(LoadFile):
    def fileDialog(parent: any):
        dialog = QFileDialog.getOpenFileNames(filter="TIF Files (*.tif *.tiff)")
        if dialog is None or len(dialog[0]) == 0:
            return None
        else:
            return dialog[0]

    def loadData(): ...
