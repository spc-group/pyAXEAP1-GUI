# :author: Alexander Berno
"""RXES Window"""

import pyqtgraph as pg
import sys
from PyQt6 import QtWidgets, QtCore, QtGui
from matplotlib.figure import Figure
from matplotlib.backends.backend_qtagg import (
    FigureCanvasQTAgg as FigureCanvas,
    # NavigationToolbar2QT as NavigationToolbar,
)
import matplotlib
import numpy as np
from openpyxl import Workbook as ExWorkbook
from openpyxl.utils import get_column_letter as getColumnLetter
from axeap import core

matplotlib.use("QtAgg")

from ExitDialogWindow import exitDialog
from RXESSpectrumClass import Dataset, Spectrum
from FileLoad import LoadTifSpectraData, LoadH5Data
from BaseWindow import Window
from ErrorWindow import ErrorWindow
from spectraFunctions import calcDataForSpectra, calcSpectra
from LoadingBarWindow import LoadingBarWindow
from FileLoad import LoadInfoData
from SettingsWindow import SettingsWindow


AlignFlag = QtCore.Qt.AlignmentFlag


# This is to remove Qt warning messages (for parts that are known to not be problems)
def handler(msg_type, msg_log, msg_string):
    pass


# finds closest value to a given value (for transfer energy)
def find_closest(lst, k):
    closest_num = lst[0]
    for num in lst:
        if abs(num - k) < abs(closest_num - k):
            closest_num = num
        if num > k:
            break
    return closest_num


# see handler above
QtCore.qInstallMessageHandler(handler)


class Mpl3dCanvas(FigureCanvas):
    """canvas for 3d figure"""

    def __init__(self, parent=None, width=3, height=3, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(projection="3d")
        super(Mpl3dCanvas, self).__init__(fig)


class Mpl2dCanvas(FigureCanvas):
    """canvas for 2d figure"""

    def __init__(self, parent=None, width=3, height=3, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)
        super(Mpl2dCanvas, self).__init__(fig)


class RXESWindow(Window):
    """Window for viewing RXES data"""

    def __init__(self, parent: QtWidgets.QMainWindow | None, *args, **kwargs):
        super(RXESWindow, self).__init__(*args, **kwargs)

        # Default Values
        self.setWindowTitle("RXES (RIXS)")
        self.setFixedSize(800, 720)
        self.parent = parent
        if self.parent is None:
            self.no_close_dialog = True
        else:
            self.no_close_dialog = False
        self.emaps = []
        self.info_file = None
        self.i0 = None
        self.incident_energy = None
        self.normalize = False
        self.use_log = False
        self.transfer = False
        self.ela_remove = False
        self.foldernames = []
        self.datasets = []
        self.old_2d = {
            "data": [],
            "use": {
                "tr": self.transfer,
                "norm": self.normalize,
                "ela": self.ela_remove,
                "log": self.use_log,
            },
            "col_mode": None,
            "colmap_type": None,
        }
        self.data_changed = False

        # energy map assignment (if parent has an energy map)
        if self.parent is None:
            self.emap = None
        else:
            try:
                self.emap = self.parent.emap
                self.emaps.append(self.emap)
            except Exception:
                self.emap = None

        # 3D canvas init
        self.sc3d = Mpl3dCanvas(self)
        # toolbar = NavigationToolbar(self.sc3d, self)
        self.ax3d = self.sc3d.axes
        self.fixax3d()

        # 2D canvas init (contour map)
        self.sc2d = Mpl2dCanvas(self)
        self.ax2d = self.sc2d.axes
        self.ax2d.set_position((0.23, 0.16, 0.73, 0.8))
        self.fixax2d()
        # self.sc2d = pg.plot()
        # # self.sc2d.setBackground("w")
        # self.sc2d.plotItem.getAxis("left").setLabel(text="Emission", **label_style)
        # self.sc2d.plotItem.getAxis("bottom").setLabel(text="Incident", **label_style)
        # self.sc2d.setFixedSize(300, 300)
        # self.ax2d = pg.ScatterPlotItem()
        # self.sc2d.addItem(self.ax2d)

        # Menu Bar
        menubar = QtWidgets.QMenuBar()
        self.setMenuBar(menubar)

        # File Menu
        filemenu = QtWidgets.QMenu("Save", menubar)

        # adding the file menu to the menu bar
        menubar.addAction(filemenu.menuAction())
        menubar.setGeometry(0, 0, 800, 22)
        filemenu.setFixedSize(240, 136)

        # Save all spectra "button"
        self.save_all_button = QtGui.QAction("Save Selected Datasets As...")
        self.save_all_button.triggered.connect(self.saveAllSpectra)
        self.save_all_button.setDisabled(True)
        self.save_all_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_all_button)
        filemenu.addSeparator()

        # Save selected spectra "button"
        self.save_disp_button = QtGui.QAction("Save Averaged Spectra As...")
        self.save_disp_button.triggered.connect(self.saveDispSpectra)
        self.save_disp_button.setDisabled(True)
        self.save_disp_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_disp_button)
        filemenu.addSeparator()

        # Save surface plot image
        self.save_surf_button = QtGui.QAction("Save Surface Plot Figure As...")
        self.save_surf_button.triggered.connect(self.saveSurfFigure)
        self.save_surf_button.setDisabled(True)
        self.save_surf_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_surf_button)
        filemenu.addSeparator()

        # Save contour plot image
        self.save_cont_button = QtGui.QAction("Save Contour Map Figure As...")
        self.save_cont_button.triggered.connect(self.saveContFigure)
        self.save_cont_button.setDisabled(True)
        self.save_cont_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_cont_button)
        filemenu.addSeparator()

        # Save emission slice "button"
        self.save_em_button = QtGui.QAction("Save Emission Slice As...")
        self.save_em_button.triggered.connect(self.saveEmissionSlice)
        self.save_em_button.setDisabled(True)
        self.save_em_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_em_button)
        filemenu.addSeparator()

        # Save incident slice "button"
        self.save_inc_button = QtGui.QAction("Save Incident Slice As...")
        self.save_inc_button.triggered.connect(self.saveIncidentSlice)
        self.save_inc_button.setDisabled(True)
        self.save_inc_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_inc_button)

        # Emission canvas init
        label_style = {"color": "#444", "font-size": "14pt"}
        self.emsc = pg.plot()
        self.emsc.setBackground("w")
        self.emsc.plotItem.getAxis("left").setLabel(text="Signal Counts", **label_style)
        self.emsc.plotItem.getAxis("bottom").setLabel(
            text="Incident (eV)", **label_style
        )
        self.emsc.setFixedSize(300, 300)

        # Incident canvas init
        self.incsc = pg.plot()
        self.incsc.setBackground("w")
        self.incsc.plotItem.getAxis("left").setLabel(
            text="Signal Counts", **label_style
        )
        self.incsc.plotItem.getAxis("bottom").setLabel(
            text="Emission (eV)", **label_style
        )
        self.incsc.setFixedSize(300, 300)

        # RXES data button
        load_rxes_button = QtWidgets.QPushButton("Load RXES Data...")
        load_rxes_button.clicked.connect(self.loadRXES)
        load_rxes_button.setFixedSize(160, 30)

        # Refresh button
        self.refresh_button = QtWidgets.QPushButton("Refresh")
        self.refresh_button.setDisabled(True)
        self.refresh_button.clicked.connect(self.refresh)
        self.refresh_button.setFixedSize(160, 30)

        # Energy map selection box
        self.emap_combo = QtWidgets.QComboBox()
        self.emap_combo.setFixedSize(140, 30)
        if self.emap is not None:
            self.emap_combo.addItem(self.emap.name)

        # Load Energy Map button
        emap_load_button = QtWidgets.QPushButton("Load Energy Map...")
        emap_load_button.setFixedSize(140, 30)
        emap_load_button.clicked.connect(self.loadEmap)

        # "norm" grid, not only used for normalization
        norm_area = QtWidgets.QScrollArea()
        norm_widget = QtWidgets.QWidget()
        norm_grid = QtWidgets.QGridLayout(norm_widget)

        # Normalize checkbox
        norm_check = QtWidgets.QCheckBox("Normalize")
        norm_check.stateChanged.connect(self.normSwitch)

        # Load Info File button
        info_load_button = QtWidgets.QPushButton("Load Info File...")
        info_load_button.setFixedSize(120, 30)
        info_load_button.clicked.connect(self.loadInfoFile)

        # colour mode selection box and label
        try:
            cmap = SettingsWindow.getFileSettings()["cmap"]
        except Exception:
            cmap = "pcolor"
        self.colour_mode = QtWidgets.QComboBox()
        self.colour_mode.setFixedSize(74, 30)
        self.colour_mode.addItem("PColor", "pcolor")
        self.colour_mode.addItem("Contour", "contour")
        if cmap == "pcolor":
            self.colour_mode.setCurrentIndex(0)
        else:
            self.colour_mode.setCurrentIndex(1)
        colour_mode_label = QtWidgets.QLabel("2D Mode:")

        # use log for intensity checkbox
        log_check = QtWidgets.QCheckBox("Log Intensity")
        log_check.stateChanged.connect(self.logSwitch)

        # "transfer energy" checkbox
        transfer_check = QtWidgets.QCheckBox("Transfer Energy")
        transfer_check.stateChanged.connect(self.transferSwitch)

        # "elastic removal" checkbox
        ela_check = QtWidgets.QCheckBox("Ela. Removal")
        ela_check.stateChanged.connect(self.elaSwitch)
        ela_check.setToolTip(
            "Elastic Removal. Removes peaks where Incident equals Emission."
        )

        # number of points to include (squared)
        num_points_label = QtWidgets.QLabel("Rows/Cols:")
        self.num_points = QtWidgets.QLineEdit()
        self.num_points.setValidator(QtGui.QIntValidator(1, 9999))
        self.num_points.setText("50")
        self.num_points.setFixedWidth(64)

        # colour map type
        colmap_label = QtWidgets.QLabel("Map:")
        self.colmap_type = QtWidgets.QComboBox()
        self.colmap_type.setFixedSize(74, 30)
        self.colmap_type.addItem("Viridis", "viridis")
        self.colmap_type.addItem("Plasma", "plasma")
        self.colmap_type.addItem("Inferno", "inferno")
        self.colmap_type.addItem("Magma", "magma")
        self.colmap_type.addItem("Cividis", "cividis")
        self.colmap_type.addItem("Gray", "gray")

        # add everything to norm_area
        norm_grid.addWidget(info_load_button, 0, 0, 1, 2)
        norm_grid.addWidget(colour_mode_label, 1, 0)
        norm_grid.addWidget(self.colour_mode, 1, 1)
        norm_grid.addWidget(norm_check, 2, 0, 1, 2)
        norm_grid.addWidget(log_check, 3, 0, 1, 2)
        norm_grid.addWidget(ela_check, 4, 0, 1, 2)
        norm_grid.addWidget(transfer_check, 5, 0, 1, 2)
        norm_grid.addWidget(num_points_label, 6, 0)
        norm_grid.addWidget(self.num_points, 6, 1)
        norm_grid.addWidget(colmap_label, 7, 0)
        norm_grid.addWidget(self.colmap_type, 7, 1)
        norm_area.setWidget(norm_widget)

        # Emission and Incident selection
        label_em = QtWidgets.QLabel("Emission")
        self.select_em = QtWidgets.QLineEdit()
        self.select_em.setDisabled(True)
        label_inc = QtWidgets.QLabel("Incident")
        self.select_inc = QtWidgets.QLineEdit()
        self.select_inc.setDisabled(True)
        self.em_inc_button = QtWidgets.QPushButton("Plot")
        self.em_inc_button.clicked.connect(self.calcEmInc)
        self.em_inc_button.setDisabled(True)

        # Connects everything to the RXES window
        widget = QtWidgets.QWidget()
        self.mlayout = QtWidgets.QGridLayout(widget)
        self.mlayout.addWidget(self.refresh_button, 0, 0, AlignFlag.AlignHCenter)
        self.mlayout.addWidget(load_rxes_button, 1, 0, AlignFlag.AlignHCenter)
        self.mlayout.addWidget(self.emap_combo, 0, 5, AlignFlag.AlignRight)
        # self.mlayout.addWidget(toolbar, 1, 1, AlignFlag.AlignLeft)
        self.mlayout.addWidget(emap_load_button, 1, 5, AlignFlag.AlignRight)
        self.mlayout.addWidget(norm_area, 2, 0)
        self.mlayout.addWidget(self.sc3d, 2, 1, 1, 4, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.sc2d, 2, 2, 1, 4, AlignFlag.AlignRight)
        self.mlayout.addWidget(label_em, 3, 1, AlignFlag.AlignRight)
        self.mlayout.addWidget(self.select_em, 3, 2)
        self.mlayout.addWidget(label_inc, 3, 3)
        self.mlayout.addWidget(self.select_inc, 3, 4)
        self.mlayout.addWidget(self.em_inc_button, 3, 5)
        self.mlayout.addWidget(self.emsc, 4, 1, 1, 4, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.incsc, 4, 2, 1, 4, AlignFlag.AlignRight)
        self.mlayout.setColumnMinimumWidth(0, 170)
        self.mlayout.setColumnMinimumWidth(2, 180)
        self.mlayout.setColumnMinimumWidth(4, 180)

        self.setCentralWidget(widget)

        self.show()

    # Sets labels and view angle for 3D graph
    def fixax3d(self):
        self.ax3d.set_xlabel("Incident (eV)")
        self.ax3d.set_ylabel("Emission (eV)")
        # self.ax3d.set_zlabel("Intensity")
        self.ax3d.view_init(25, 225, 0)

    def fixax3dtr(self):
        self.ax3d.set_xlabel("Incident (eV)")
        self.ax3d.set_ylabel("Transfer")
        # self.ax3d.set_zlabel("Intensity")
        self.ax3d.view_init(25, 225, 0)

    # Sets labels for contour map
    def fixax2d(self):
        self.ax2d.set_xlabel("Incident (eV)")
        self.ax2d.set_ylabel("Emission (eV)")

    def fixax2dtr(self):
        self.ax2d.set_xlabel("Incident (eV)")
        self.ax2d.set_ylabel("Transfer")

    # flips whether or not to normalize
    def normSwitch(self):
        self.normalize = not self.normalize

    # flips whether or not to log intensities
    def logSwitch(self):
        self.use_log = not self.use_log

    # sets whether or not to use transfer energy
    def transferSwitch(self):
        self.transfer = not self.transfer

    # sets whether or not to use elastic removal
    def elaSwitch(self):
        self.ela_remove = not self.ela_remove

    # handles close event so a confirmation window can appear
    def closeEvent(self, event):
        # The no_close_dialog exists so the window can be closed by a MainWindow with no issue
        if not self.no_close_dialog:
            if hasattr(self.parent, "confirm_on_close"):
                if self.parent.confirm_on_close:
                    confirm = exitDialog(self)
                else:
                    confirm = True
            else:
                confirm = exitDialog(self)
            if confirm:
                if hasattr(self.parent, "childWindow"):
                    self.parent.childWindow = None
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    # loads information file
    def loadInfoFile(self):
        directory = LoadInfoData.fileDialog(self)
        if directory[0] == "":
            return
        self.info_file, self.info_table = LoadInfoData.loadData(
            self, directory, rtype="rxes"
        )
        try:
            self.incident_energy = self.info_table["Energy"]
            self.i0 = self.info_table["I0"]
        except Exception:
            self.incident_energy = self.info_table[0]
            self.i0 = self.info_table[1]
        self.data_changed = True

    # Main function for loading RXES data
    def loadRXES(self):

        if len(self.emaps):
            self.emap = self.emaps[int(self.emap_combo.currentIndex() / 2)]
        elif self.emap is None:
            try:
                self.emap = self.parent.emap
            except Exception:
                self.error = ErrorWindow("XESemap")
                return

        emap = self.emap
        data = calcDataForSpectra(emap)
        dtype = self.loadType()
        if dtype == "tif":
            self.filenames = LoadTifSpectraData.fileDialog(self)
        elif dtype == "h5py":
            self.filenames = LoadH5Data.fileDialog(self)
        else:
            raise TypeError(f"Unknown dtype {dtype}")

        if not self.filenames:
            return

        scanset = []
        energies = []
        i0s = []
        if dtype == "h5py":
            for l, j in enumerate(self.filenames, 1):
                images, energy, i0 = LoadH5Data.loadData(j)
                temp_scans = [core.Scan(np.swapaxes(img, 0, 1)) for img in images]
                energies.append(energy)
                i0s.append(i0)
                LoadWindow = LoadingBarWindow(
                    f"Loading RXES (RIXS) data... ({l}/{len(self.filenames)})",
                    len(temp_scans),
                )
                for i in temp_scans:
                    if LoadWindow.wasCanceled():
                        break
                    spectra, _, _ = calcSpectra(i, emap, data, dtype)
                    scanset.append(spectra)
                    LoadWindow.add()
                    QtWidgets.QApplication.processEvents()

        else:
            LoadWindow = LoadingBarWindow(
                "Loading RXES (RIXS) data...", len(self.filenames)
            )
            for i in self.filenames:
                if LoadWindow.wasCanceled():
                    break
                spectra, energy, i0 = calcSpectra(i, emap, data, dtype)
                if energy and len(energy) == len(spectra):
                    # self.incident_energy = energy
                    # self.i0 = i0
                    energies.append(energy)
                    i0s.append(i0)
                if type(spectra) is list:
                    scanset += spectra
                else:
                    scanset.append(spectra)
                LoadWindow.add()
                QtWidgets.QApplication.processEvents()

        LoadWindow.deleteLater()

        if LoadWindow.wasCanceled():
            return

        dname = self.filenames[0]
        if dtype == "tif":
            dname = dname[: dname.rfind("/")]
        dname = dname[dname.rfind("/") + 1 :]
        self.foldernames.append(dname)
        if dtype == "h5py":
            dataset = Dataset(
                self,
                dname,
                scanset,
                len(self.datasets),
                energy=energy,
                i0=i0,
                enabled=True,
            )
        else:
            dataset = Dataset(self, dname, scanset, len(self.datasets))
        self.datasets.append(dataset)
        self.addDataCheckbox()

        failed = self.setData()  # returns True if failed, otherwise None
        if failed:
            self.foldernames.remove(dname)
            self.datasets.remove(dataset)
            self.addDataCheckbox()
            return

        self.data_changed = True

        self.setSubLimits()
        self.graph3dSpectra()
        self.graph2dSpectra()

    def setSubLimits(self):
        mininte = 100000
        maxinte = 0
        minem = 100000
        maxem = 0
        mininc = 100000
        maxinc = 0
        for s in self.spectra:
            if max(s.inte) > maxinte:
                maxinte = max(s.inte)
            if min(s.inte) < mininte:
                mininte = min(s.inte)
            if max(s.em) > maxem:
                maxem = max(s.em)
            if min(s.em) < minem:
                minem = min(s.em)
            if max(s.inc) > maxinc:
                maxinc = max(s.inc)
            if min(s.inc) < mininc:
                mininc = min(s.inc)
        # self.select_em.setValidator(QtGui.QIntValidator(int(minem), int(maxem)))
        # self.select_inc.setValidator(QtGui.QIntValidator(int(mininc), int(maxinc)))
        regexp = QtCore.QRegularExpression("((\d+|\d+-\d+),?)*")
        self.select_em.setValidator(QtGui.QRegularExpressionValidator(regexp))
        self.select_inc.setValidator(QtGui.QRegularExpressionValidator(regexp))
        self.em_limits = (minem, maxem)
        self.inc_limits = (mininc, maxinc)
        self.select_em.setDisabled(False)
        self.select_inc.setDisabled(False)
        self.em_inc_button.setDisabled(False)

    # sets spectra data, including creating Spectrum classes
    def setData(self, scanset=None, do_return=False):

        if scanset is None:
            new_set = []
            for dataset in self.datasets:
                if dataset.enabled:
                    new_set.append((dataset.data, dataset.energy, dataset.i0))

            scanset = new_set
        if not len(scanset):
            if not do_return:
                self.spectra = []
            return

        ul = self.use_log
        tr = self.transfer
        ela = self.ela_remove
        if not self.normalize and (
            self.incident_energy is None and any(i[1] is None for i in scanset)
        ):
            # if not multi:
            #     self.spectra = [
            #         Spectrum(self, s, i, ul=ul, tr=tr, ela=ela)
            #         for i, s in enumerate(scanset)
            #     ]
            # else:
            spectra = []
            for i, _ in enumerate(scanset[0][0]):
                s = [scanset[j][0][i] for j, _ in enumerate(scanset)]
                spectra.append(Spectrum(self, s, i, ul=ul, tr=tr, ela=ela))

        elif (self.i0 is None and any(i[2] is None for i in scanset)) or (
            self.incident_energy is None and any(i[1] is None for i in scanset)
        ):
            self.error = ErrorWindow("noInfoRXES")
            return True

        else:
            # if multi:
            slen = len(scanset[0][0])
            # else:
            #     slen = len(scanset)

            i0inc = False
            if self.i0 is not None and self.incident_energy is not None:
                i0inc = True
                inc = self.incident_energy
                i0 = self.i0

                inclen = len(inc)
                i0len = len(i0)

            if self.normalize:
                if i0inc:
                    if slen != i0len or slen != inclen or i0len != inclen:
                        self.error = ErrorWindow("NotEnoughData")
                        return True
                    spectra = []
                    for i, _ in enumerate(scanset[0][0]):
                        s = [scanset[j][0][i] for j, _ in enumerate(scanset)]
                        spectra.append(
                            Spectrum(self, s, i, inc[i], i0[i], ul=ul, tr=tr, ela=ela)
                        )
                else:
                    spectra = []
                    for i, _ in enumerate(scanset[0][0]):
                        s = [scanset[j][0][i] for j, _ in enumerate(scanset)]
                        inc = [scanset[j][1][i] for j, _ in enumerate(scanset)]
                        i0 = [scanset[j][2][i] for j, _ in enumerate(scanset)]
                        spectra.append(
                            Spectrum(self, s, i, inc, i0, ul=ul, tr=tr, ela=ela)
                        )

            else:
                # if not multi:
                #     self.spectra = [
                #         Spectrum(self, s, i, inc[i], ul=ul, tr=tr, ela=ela)
                #         for i, s in enumerate(scanset)
                #     ]
                # else:
                if i0inc:
                    spectra = []
                    for i, _ in enumerate(scanset[0][0]):
                        s = [scanset[j][0][i] for j, _ in enumerate(scanset)]
                        spectra.append(
                            Spectrum(self, s, i, inc[i], ul=ul, tr=tr, ela=ela)
                        )
                else:
                    spectra = []
                    for i, _ in enumerate(scanset[0][0]):
                        s = [scanset[j][0][i] for j, _ in enumerate(scanset)]
                        inc = [scanset[j][1][i] for j, _ in enumerate(scanset)]
                        spectra.append(Spectrum(self, s, i, inc, ul=ul, tr=tr, ela=ela))

        if do_return:
            return spectra
        else:
            self.spectra = spectra

    # creates the checkbox layout (and recreates it to fix formatting)
    def addDataCheckbox(self):

        try:
            self.mlayout.removeWidget(self.checks)
        except Exception:
            pass

        self.checks = QtWidgets.QScrollArea()
        self.checks_widget = QtWidgets.QWidget()
        self.checks_grid = QtWidgets.QGridLayout(self.checks_widget)

        self.datasets = [
            Dataset(self, d.name, d.data, d.num, d.energy, d.i0, d.enabled)
            for d in self.datasets
        ]
        for dataset in self.datasets:
            self.checks_grid.addWidget(dataset.box, dataset.num, 0)
            self.checks_grid.setRowMinimumHeight(dataset.num, 20)

        self.checks.setWidget(self.checks_widget)
        self.mlayout.addWidget(self.checks, 4, 0)

    # sets data and graphs data all in one (for simpler calling)
    def refresh(self):
        failed = self.setData()
        if not failed:
            self.graph3dSpectra()
            self.graph2dSpectra()
            self.setSubLimits()

    # This is the 3d graph
    def graph3dSpectra(self):
        self.ax3d.clear()
        if self.transfer:
            self.fixax3dtr()
        else:
            self.fixax3d()

        self.refresh_button.setDisabled(False)

        if not len(self.spectra):
            self.sc3d.draw_idle()
            return

        x, y, z = [], [], []
        for i, s in enumerate(self.spectra):

            if self.transfer and i < len(self.spectra) - 1:
                if s.em[-1] > self.spectra[i + 1].em[-1]:
                    continue
            x.append(s.inc)
            y.append(s.em)
            z.append(s.inte)
        x = np.asarray(x)
        y = np.asarray(y)
        z = np.asarray(z)

        count = int(self.num_points.text())
        self.ax3d.plot_surface(
            x,
            y,
            z,
            cmap=self.colmap_type.currentData(),
            rcount=count,
            ccount=count,
            antialiased=False,
        )
        self.sc3d.draw_idle()

    # This is the contour map
    def graph2dSpectra(self, forced=False):
        # skips drawing 2d if it hasn't changed
        if (
            not (
                self.data_changed
                or self.transfer != self.old_2d["use"]["tr"]
                or self.normalize != self.old_2d["use"]["norm"]
                or self.ela_remove != self.old_2d["use"]["ela"]
                or self.use_log != self.old_2d["use"]["log"]
                or self.colour_mode.currentData() != self.old_2d["col_mode"]
                or self.colmap_type.currentData() != self.old_2d["colmap_type"]
            )
            and not forced
        ):
            return

        # min and max values used for analysis later
        # these values are used in RXESWindow.calcEmInc

        if not len(self.spectra):
            self.ax2d.cla()
            self.fixax2d()
            self.sc2d.draw_idle()
            return

        x, y, z = [], [], []
        for i, s in enumerate(self.spectra):
            if self.transfer and i < len(self.spectra) - 1:
                if s.em[-1] > self.spectra[i + 1].em[-1]:
                    continue
            x.append(s.inc)
            y.append(s.em)
            z.append(s.inte)
        x = np.asarray(x)
        y = np.asarray(y)
        z = np.asarray(z)

        self.data_changed = False

        self.ax2d.cla()
        if self.transfer:
            self.fixax2dtr()
        else:
            self.fixax2d()
        col_mode = self.colour_mode.currentData()
        colmap = self.colmap_type.currentData()
        if col_mode == "contour":
            self.ax2d.contourf(x, y, z, levels=10, extend="both", cmap=colmap)
        elif col_mode == "pcolor":
            self.ax2d.pcolor(x, y, z, cmap=colmap)

        self.old_2d = {
            "data": [x, y, z],
            "use": {
                "tr": self.transfer,
                "norm": self.normalize,
                "ela": self.ela_remove,
                "log": self.use_log,
            },
            "col_mode": col_mode,
            "colmap_type": colmap,
        }

        self.save_all_button.setDisabled(False)
        self.save_disp_button.setDisabled(False)
        self.save_surf_button.setDisabled(False)
        self.save_cont_button.setDisabled(False)
        self.sc2d.draw_idle()

        self.save_em_button.setDisabled(True)
        self.save_inc_button.setDisabled(True)
        self.select_em.setText("")
        self.select_inc.setText("")
        self.emsc.plotItem.clear()
        self.incsc.plotItem.clear()

    # Get datapoints for Emission and Incident vs Intensity 2D graphs
    def calcEmInc(self):
        inc = self.select_inc.text().split(",")
        em = self.select_em.text().split(",")
        if not (inc or em):
            return

        if inc:
            inc_data = []
            for i in inc:
                current = float(i)
                mininc, maxinc = self.inc_limits
                if not (mininc <= current <= maxinc):
                    print(current, maxinc, mininc)
                    self.error = ErrorWindow("invalidEmIncRXES")
                    return
                del mininc, maxinc

                # Incident Calc
                try:
                    inc_spectrum = self.spectra[current]
                except Exception:
                    ls = None
                    for s in self.spectra:
                        if s.inc[0] > current:
                            if ls is None:
                                ls = s
                            break
                        ls = s

                    if min(abs(ls.inc[0] - current), abs(s.inc[0] - current)) == abs(
                        s.inc[0] - current
                    ):
                        inc_spectrum = s
                    else:
                        inc_spectrum = ls

                inc_emission = inc_spectrum.em
                inc_intensity = inc_spectrum.inte
                inc_data.append((inc_emission, inc_intensity))

        elif not inc:
            inc_data = []

        if em:
            em_data = []
            for i in em:
                current = float(i)
                minem, maxem = self.em_limits
                if not (minem <= current <= maxem):
                    print(current, maxem, minem)
                    self.error = ErrorWindow("invalidEmIncRXES")
                    return
                del minem, maxem

                # Emission Calc
                em_incident = []
                em_intensity = []
                for s in self.spectra:
                    for j, e in enumerate(s.em):
                        if current - 0.1 <= e <= current + 0.1:
                            break
                    em_incident.append(s.inc[0])
                    em_intensity.append(s.inte[j])
                em_data.append((em_incident, em_intensity))

        elif not em:
            em_data = []

        self.graphEmInc(em_data, inc_data)
        self.save_em_button.setDisabled(False)
        self.save_inc_button.setDisabled(False)

    # Graph Emission and Incident vs Intensity 2D graphs
    def graphEmInc(self, em, inc):
        if em:
            self.emsc.plotItem.clear()
            for e in em:
                self.emsc.plotItem.plot(e[0], e[1], pen=pg.mkPen(color="k", width=2))
        if inc:
            self.incsc.plotItem.clear()
            for i in inc:
                self.incsc.plotItem.plot(i[0], i[1], pen=pg.mkPen(color="k", width=2))

    def saveSpectra(self, spectra=None):
        if spectra is None:
            spectra = self.spectra

        dialog = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Spectra",
            filter=("Excel Spreadsheet (*.xlsx)\nSimple Text Layout (*.csv)"),
        )

        if type(spectra[0]) is Spectrum:
            if dialog[1] == "Excel Spreadsheet (*.xlsx)":
                wb = ExWorkbook()
                ws = wb.active

                lines = [[] for _, _ in enumerate(spectra[0].inc)]
                lines.append([])
                lines.append([])
                lines.append([])
                for spect in spectra:
                    lines[0].append(f"Spectrum {spect.num}")
                    lines[0].append("")
                    lines[0].append("")
                    lines[0].append("")
                    lines[1].append("Incident Energy (eV)")
                    lines[1].append("Emission Energy (eV)")
                    lines[1].append("Signal Counts")
                    lines[1].append("")
                    texts = [
                        [str(spect.inc[j]), str(spect.em[j]), str(spect.inte[j]), ""]
                        for j, _ in enumerate(spect.inc)
                    ]
                    for k, item in enumerate(texts):
                        for l in range(4):
                            lines[k + 2].append(item[l])
                for i, line in enumerate(lines, 1):
                    for j, item in enumerate(line, 1):
                        try:
                            n = float(item)
                        except Exception:
                            n = item
                        ws.cell(i, j + 1).value = n
                for i in range(int(len(lines[1]) / 4)):
                    ws.merge_cells(
                        start_row=1,
                        end_row=1,
                        start_column=i * 4 + 2,
                        end_column=i * 4 + 4,
                    )
                    ws.column_dimensions[getColumnLetter(i * 4 + 2)].width = 18
                    ws.column_dimensions[getColumnLetter(i * 4 + 3)].width = 18
                    ws.column_dimensions[getColumnLetter(i * 4 + 4)].width = 8

                ws.cell(1, 1).value = f"Normalized: {self.normalize}"
                ws.cell(2, 1).value = f"Logged Intensity: {self.use_log}"
                ws.cell(3, 1).value = f"Elastic Removal: {self.ela_remove}"
                ws.cell(4, 1).value = f"Transfer Energy: {self.transfer}"
                ws.column_dimensions[getColumnLetter(1)].width = 24

                wb.save(dialog[0])
                wb.close()

            elif dialog[1] == "Simple Text Layout (*.csv)":
                direct = open(dialog[0], "+w")
                direct.seek(0)
                direct.truncate()
                lines = ["" for _, _ in enumerate(spectra[0].inc)]
                lines.append("")
                lines.append("")
                lines[0] += f"Normalized: {self.normalize},"
                lines[1] += f"Logged Intensity: {self.use_log},"
                lines[2] += f"Elastic Removal: {self.ela_remove},"
                lines[3] += f"Transfer Energy: {self.transfer},"
                for l, _ in enumerate(lines[4:], 4):
                    lines[l] += ","
                for spect in spectra:
                    lines[0] += f"Spectrum {spect.num},,,,"
                    lines[
                        1
                    ] += "Incident Energy (eV),Emission Energy (eV), Signal Counts,,"
                    texts = [
                        f"{spect.inc[j]},{spect.em[j]},{spect.inte[j]},,"
                        for j, _ in enumerate(spect.inc)
                    ]
                    for k, string in enumerate(texts):
                        lines[k + 2] += string
                text = ""
                for line in lines:
                    text += line + "\n"
                direct.write(text)
                direct.close()
        else:
            pass

    def saveAllSpectra(self):
        data = [(d.data, d.energy, d.i0) for d in self.datasets if d.enabled]
        spectra = []
        for d in data:
            spectra.append(self.setData([d], True))
        self.saveSpectra(spectra)

    def saveDispSpectra(self):
        self.saveSpectra()

    def saveSurfFigure(self):
        dialog = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Surface Figure",
            filter=("Image (*.png)\nPDF (Vector) (*.pdf)"),
        )
        self.ax3d.get_figure().savefig(dialog[0])

    def saveContFigure(self):
        dialog = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Contour Figure",
            filter=("Image (*.png)\nPDF (Vector) (*.pdf)"),
        )
        self.ax2d.get_figure().savefig(dialog[0])

    def saveEmissionSlice(self):
        pass

    def saveIncidentSlice(self):
        pass


# creates an RXES window when file is run
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon("icons/spc-logo-nobg.png"))
    W = RXESWindow(None)
    # sys.exit(app.exec())
    app.exec()
