# :author: Alexander Berno

import axeap.core as core
import sys

from axeap.core.roi import HROI
import pathlib
from openpyxl import load_workbook
import numpy as np

from LoadingBarWindow import LoadingBarWindow
from ApproxWindow import ApproxWindow
from ErrorWindow import ErrorWindow
from GetPoints import GetPoints
from XESWindow import XESWindow
from RXESWindow import RXESWindow
from calibFunctions import loadCalib, calcEnergyMap, approximateROIs, getCoordsFromScans
from CalibFileClass import CalibFile
from SettingsWindow import SettingsWindow


from PyQt6 import QtCore, QtWidgets, QtGui
import pyqtgraph as pg


# Alignment flags are used to place items in a window.
AlignFlag = QtCore.Qt.AlignmentFlag

# default directory for file selection menus (Desktop)
desktop_directory = str(pathlib.Path.home() / "Desktop")


# Main window
class MainWindow(QtWidgets.QMainWindow):
    """Main Window for application. Allows the viewing of data."""

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)

        # Defaults
        self.sel_x1 = None
        self.sel_y1 = None
        self.sel_x2 = None
        self.sel_y2 = None
        self.sel_ax = None
        self.sel_rect = [None, []]
        self.rects = []
        self.drawn_calib = False
        self.left_button = False
        self.ax = None
        self.childWindow = None
        self.info_file = None
        self.emap_img = None
        self.calib_grid_scroll = None

        self.setWindowTitle("pyAXEAP1")
        self.setFixedSize(960, 574)

        self.sc = pg.plot()
        self.sc.setBackground("w")
        self.ax = pg.ScatterPlotItem()
        self.sc.addItem(self.ax)

        # button for opening calibration data
        cali_button = QtWidgets.QPushButton("Import Images...")
        cali_button.setFixedSize(120, 30)
        cali_button.clicked.connect(self.openPath)

        # button for opening XES menu
        xes_button = QtWidgets.QPushButton("XES")
        xes_button.clicked.connect(self.runXES)
        xes_button.setFixedSize(120, 30)

        # button for opening RXES menu
        rxes_button = QtWidgets.QPushButton("RXES")
        rxes_button.clicked.connect(self.runRXES)
        rxes_button.setFixedSize(120, 30)

        # button for loading an existing energy map
        emap_load_button = QtWidgets.QPushButton("Load Energy Map...")
        emap_load_button.setFixedSize(120, 30)
        emap_load_button.clicked.connect(self.loadEmap)

        # get settings from settings file, or load defaults
        try:
            settings = self.getSettings()
        except Exception:
            settings = self.getDefaultSettings()

        # settings button
        set_button = QtWidgets.QPushButton("Settings")
        set_button.clicked.connect(lambda: self.openSettings(settings))
        set_button.setFixedSize(120, 30)

        # energy map buttons
        emap_area = QtWidgets.QScrollArea()
        emap_area.setFixedSize(286, 170)
        emap_widget = QtWidgets.QWidget()
        emap_grid = QtWidgets.QGridLayout(emap_widget)

        mincuts_label = QtWidgets.QLabel("Minimum Cuts")
        self.mincuts = QtWidgets.QSpinBox()
        self.mincuts.setMinimumSize(128, 20)
        self.mincuts.setMaximum(10000000)
        self.mincuts.setValue(int(settings["default_min_cuts"]))
        maxcuts_label = QtWidgets.QLabel("Maximum cuts")
        self.maxcuts = QtWidgets.QSpinBox()
        self.maxcuts.setMinimumSize(128, 20)
        self.maxcuts.setMaximum(10000000)
        self.maxcuts.setValue(int(settings["default_max_cuts"]))
        self.reload_calib = QtWidgets.QPushButton("Refresh")
        self.reload_calib.clicked.connect(lambda: self.getCalibSpectra(False))
        self.reload_calib.setDisabled(True)
        self.approx_rois = QtWidgets.QPushButton("Set ROIs Automatically")
        self.approx_rois.setDisabled(True)
        self.approx_rois.clicked.connect(self.approxROIs)
        self.load_info_file_button = QtWidgets.QPushButton("Load Information File")
        self.load_info_file_button.setDisabled(True)
        self.load_info_file_button.clicked.connect(self.loadInfoFile)
        self.emap_calc_button = QtWidgets.QPushButton("Caibrate")
        self.emap_calc_button.clicked.connect(self.calcEmap)
        self.emap_calc_button.setDisabled(True)
        self.emap_save_button = QtWidgets.QPushButton("Save Energy Map As...")
        self.emap_save_button.clicked.connect(self.saveEmap)
        self.emap_save_button.setDisabled(True)

        # energy map layout connections
        emap_grid.addWidget(mincuts_label, 0, 0)
        emap_grid.addWidget(self.mincuts, 0, 1)
        emap_grid.addWidget(maxcuts_label, 1, 0)
        emap_grid.addWidget(self.maxcuts, 1, 1)
        emap_grid.addWidget(self.load_info_file_button, 2, 0)
        emap_grid.addWidget(self.reload_calib, 2, 1)
        emap_grid.addWidget(self.approx_rois, 3, 0)
        emap_grid.addWidget(self.emap_calc_button, 4, 0)
        emap_grid.addWidget(self.emap_save_button, 4, 1)
        emap_area.setWidget(emap_widget)

        # main "widget"; basically everything that is within the main window
        mwidget = QtWidgets.QWidget()

        # connects everything to the window with grid organization
        self.mlayout = QtWidgets.QGridLayout(mwidget)
        self.mlayout.addWidget(cali_button, 0, 0, AlignFlag.AlignCenter)
        self.mlayout.addWidget(emap_load_button, 0, 1, AlignFlag.AlignCenter)
        self.mlayout.addWidget(xes_button, 1, 0, AlignFlag.AlignCenter)
        self.mlayout.addWidget(rxes_button, 1, 1, AlignFlag.AlignCenter)
        self.mlayout.addWidget(set_button, 1, 2, AlignFlag.AlignLeft)
        self.mlayout.addWidget(
            emap_area, 2, 0, 1, 2, AlignFlag.AlignLeft | AlignFlag.AlignTop
        )
        self.mlayout.addWidget(self.sc, 2, 2, 2, 1, AlignFlag.AlignCenter)
        self.mlayout.setColumnMinimumWidth(2, 500)

        # shows the menu
        self.setCentralWidget(mwidget)
        self.show()

    def getSettings(self):
        with open("settings.ini", "r") as s:
            lines = s.readlines()
            s.close()
        settings = {}
        for line in lines:
            if not len(line) or line[0] == "#":
                continue
            if "\n" in line:
                line = line[: line.find("\n")]
            settings[line[: line.find(" =")]] = line[line.find("= ") + 2 :]
        return settings

    def getDefaultSettings(self):
        settings = {"default_min_cuts": "3", "default_max_cuts": "10000"}
        return settings

    def openSettings(self, settings: dict | None = None):
        if settings is None:
            try:
                settings = self.getSettings()
            except Exception:
                settings = self.getDefaultSettings()

        self.SettingsWindow = SettingsWindow(self, settings)

    def closeEvent(self, event):
        super().closeEvent(event)
        if self.childWindow is not None:
            self.childWindow.close()
        self.deleteLater()

    def openPath(self):
        self.calibfiledir = QtWidgets.QFileDialog.getOpenFileNames(
            parent=self, directory=desktop_directory, filter="TIF Files (*.tif *.tiff)"
        )

        if self.calibfiledir[0] != [] and self.calibfiledir[1] != "":
            self.calibscans = loadCalib(self.calibfiledir[0])
            self.getCalibSpectra(True)

    def loadInfoFile(self):
        old = self.info_file
        self.info_file = QtWidgets.QFileDialog.getOpenFileName(
            parent=self, directory=desktop_directory
        )
        if self.info_file[0][-5:] == ".xlsx":
            wb = load_workbook(self.info_file[0], read_only=True)
            ws = wb.worksheets[0]
            for i, scan in enumerate(self.calibscans.items):
                line = ws.cell(i + 1, 1).value
                scan.meta["IncidentEnergy"] = line
                self.calib_energies[i].changeVal(line)
            wb.close()

        elif self.info_file[1] != "":
            try:
                self.calibscans.addCalibRunInfo(core.CalibRunInfo(self.info_file[0]))
                for i, e in enumerate(self.calib_energies):
                    e.changeVal(self.calibscans.items[i].meta["IncidentEnergy"])

            except Exception or Warning:
                self.info_file = old
                self.error = ErrorWindow("badInfoFile")
                return
        else:
            self.info_file = old

    def getCalibSpectra(self, runinit: bool):
        minc = self.mincuts.value()
        maxc = self.maxcuts.value()
        if minc > maxc:
            self.error = ErrorWindow("minmaxcuts")
            return
        self.LoadWindow = LoadingBarWindow(
            "Loading calibration data...", len(self.calibscans)
        )
        self.points = []
        for i in self.calibscans:
            self.points.append(getCoordsFromScans(i, reorder=True, cuts=(minc, maxc)))
            self.LoadWindow.add()
            QtWidgets.QApplication.processEvents()
        if runinit:
            self.initDrawCalibSpectra()
        else:
            self.drawCalibSpectra()

    def initDrawCalibSpectra(self):

        self.drawCalibSpectra()

        if self.calib_grid_scroll is not None:
            self.mlayout.removeWidget(self.calib_grid_scroll)
        self.calib_grid_scroll = QtWidgets.QScrollArea()
        self.calib_widget = QtWidgets.QWidget()
        self.calib_grid = QtWidgets.QGridLayout(self.calib_widget)
        self.calib_grid.addWidget(
            QtWidgets.QLabel("File Name"), 0, 0, AlignFlag.AlignLeft
        )
        self.calib_grid.addWidget(QtWidgets.QLabel("Energy"), 0, 1, AlignFlag.AlignLeft)
        self.calib_energies = [
            CalibFile(self, self.calibfiledir[0][i], i + 1)
            for i, _ in enumerate(self.calibfiledir[0])
        ]
        self.calib_grid_scroll.setWidget(self.calib_widget)
        self.calib_grid_scroll.setFixedWidth(290)
        self.mlayout.addWidget(self.calib_grid_scroll, 3, 0, 1, 2, AlignFlag.AlignLeft)

    def drawCalibSpectra(self):

        if self.ax is not None:
            self.ax.clear()
        else:
            self.ax = pg.ScatterPlotItem()

        self.rects = []
        for item in self.sc.items():
            self.sc.removeItem(item)
        self.sc.addItem(self.ax)
        self.sc.setBackground("w")

        for i in self.points:
            self.ax.addPoints(i[0], i[1], size=1, brush=(0, 0, 0, 255))

        self.drawn_calib = True
        self.reload_calib.setDisabled(False)
        self.approx_rois.setDisabled(False)
        self.load_info_file_button.setDisabled(False)

    def approxROIs(self):
        self.ApproxWindow = ApproxWindow(self)
        self.ApproxWindow.finished.connect(self.doApproxROIs)

    def doApproxROIs(self):
        if self.ApproxWindow.value is None:
            return

        self.emap_calc_button.setDisabled(False)

        numcrystals = self.ApproxWindow.value
        lencalibs = len(self.calibscans.items)
        hrois, vrois = approximateROIs(
            numcrystals,
            self.mincuts.value(),
            self.maxcuts.value(),
            self.calibscans.items[int(lencalibs / 2)],
            self.points,
        )

        self.drawCalibSpectra()

        self.rects = []
        for i, h in enumerate(hrois):
            v = vrois[i]
            rect = pg.RectROI(
                pos=(h[0], v[0]),
                size=(h[1] - h[0], v[1] - v[0]),
                pen=(255, 0, 0, 255),
                hoverPen=(255, 20, 20, 255),
                handlePen=(255, 0, 0, 255),
                handleHoverPen=(255, 0, 0, 255),
                resizable=True,
                removable=True,
            )
            rect.sigRemoveRequested.connect(self.removeROI)
            rect.addScaleHandle(pos=(0, 0), center=(1, 1))
            rect.addScaleHandle(pos=(0, 1), center=(1, 0))
            rect.addScaleHandle(pos=(1, 0), center=(0, 1))
            rect.addScaleHandle(pos=(0, 0.5), center=(1, 0.5))
            rect.addScaleHandle(pos=(1, 0.5), center=(0, 0.5))
            rect.addScaleHandle(pos=(0.5, 0), center=(0.5, 1))
            rect.addScaleHandle(pos=(0.5, 1), center=(0.5, 0))

            self.sc.addItem(rect)
            self.rects.append(rect)

    def removeROI(self, event):
        self.rects.remove(event)
        self.sc.removeItem(event)

    def calcHrois(self):
        # this section calculates the hrois
        self.hrois = []
        for i in self.rects:
            x = i.pos()[0]
            w = i.size()[0]
            self.hrois.append(HROI(x, x + w))
        # print("done")

    def calcEmap(self):
        if self.info_file is None and any(e.getVal() == 0 for e in self.calib_energies):
            self.error = ErrorWindow("noInfo")
            return
        try:
            scans = self.calibscans.items
            for i, s in enumerate(scans):
                s.meta["IncidentEnergy"] = self.calib_energies[i].getVal()
        except Exception:
            self.error = ErrorWindow("emapCalib")
            return
        points = self.points.copy()
        indexes = []
        for i, _ in enumerate(scans):
            if len(points[i][0]) < 2:
                indexes.insert(0, i)
        for i in indexes:
            scans.pop(i)
            points.pop(i)
        del indexes
        scans = core.ScanSet(scans)
        self.approx_rois.setDisabled(True)
        self.load_info_file_button.setDisabled(True)
        # self.emap_select_button.setDisabled(True)
        self.emap_calc_button.setDisabled(True)
        self.emap_save_button.setDisabled(False)
        self.calcHrois()
        try:
            self.emap = core.calcEMap(self.calibscans, self.hrois)
        except Exception:
            self.emap = calcEnergyMap(scans, points, self.hrois)
        self.drawEmap()

    def drawEmap(self):
        if self.ax is None:
            self.ax = pg.ScatterPlotItem()
        self.ax.clear()
        for i in self.rects:
            self.sc.removeItem(i)

        if self.emap_img is not None:
            self.sc.removeItem(self.emap_img)
        self.sc.setBackground("k")
        self.emap_img = pg.ImageItem(np.log(self.emap.values))
        self.sc.addItem(self.emap_img)

    def saveEmap(self):
        if self.emap is None:
            return
        dialog = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save Energy Map",
            filter="Numpy Array (*.npy)",
            directory=desktop_directory,
        )
        dir = dialog[0]
        if not len(dir):
            return
        self.emap.saveToPath(dir)
        name = dir[dir.rfind("/") + 1 :]
        self.emap.name = name[: name.rfind(".")]

    def loadEmap(self):

        text = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Open Energy Map",
            directory=str(desktop_directory),
            filter="Numpy Array (*.npy)",
        )
        if not len(text[0]):
            return
        self.emap = core.EnergyMap.loadFromPath(text[0])
        self.drawEmap()

    def runXES(self):
        if self.childWindow is not None:
            self.childWindow.close()
        self.childWindow = XESWindow(self)

    def runRXES(self):
        if self.childWindow is not None:
            self.childWindow.close()
        self.childWindow = RXESWindow(self)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon("icons/spc-logo-nobg.png"))
    W = MainWindow()
    # sys.exit(app.exec())
    app.exec()
