# :author: Alexander Berno
"""XES Window"""

from PyQt6 import QtWidgets, QtCore, QtGui
import pyqtgraph as pg
import sys

from ErrorWindow import ErrorWindow
from LoadingBarWindow import LoadingBarWindow
from spectraFunctions import calcDataForSpectra, calcSpectra
from XESSpectrumClass import Spectrum
from colourGenerator import colourGen
from ColourSelectWindow import ColourSelect
from FileLoad import LoadTifSpectraData, LoadH5Data
from ExitDialogWindow import exitDialog
from BaseWindow import Window

from openpyxl import Workbook as ExWorkbook
from openpyxl.utils import get_column_letter as getColumnLetter

AlignFlag = QtCore.Qt.AlignmentFlag


# This is to remove Qt warning messages (for parts that are known to not be problems)
def handler(msg_type, msg_log, msg_string):
    pass


# see handler above
QtCore.qInstallMessageHandler(handler)


class XESWindow(Window):
    """Window for viewing XES spectra"""

    def __init__(self, parent: QtWidgets.QMainWindow | None, *args, **kwargs):
        super(XESWindow, self).__init__(*args, **kwargs)

        # sets the main window as the parent (as well as window data)
        self.parent = parent
        self.setWindowTitle("XES")
        self.setFixedSize(960, 590)

        # XES data button
        load_xes_button = QtWidgets.QPushButton("XES Data...")
        load_xes_button.clicked.connect(self.loadXES)
        load_xes_button.setFixedSize(140, 30)

        # Canvas (graph plot)
        self.sc = pg.plot()
        self.sc.setBackground("w")
        label_style = {"color": "#444", "font-size": "14pt"}
        self.sc.plotItem.getAxis("left").setLabel(text="Signal Counts", **label_style)
        self.sc.plotItem.getAxis("bottom").setLabel(
            text="Emission Energy", **label_style
        )

        # Defaults
        if self.parent is None:
            self.no_close_dialog = True
        else:
            self.no_close_dialog = False
        self.average_spectra = None
        self.emaps = []

        # energy map assignment (if parent has an energy map)
        if self.parent is None:
            self.emap = None
        else:
            try:
                self.emap = self.parent.emap
                self.emaps.append(self.emap)
            except Exception:
                self.emap = None

        # Menu Bar
        menubar = QtWidgets.QMenuBar()
        self.setMenuBar(menubar)

        # File Menu
        filemenu = QtWidgets.QMenu("File", menubar)

        # adding the file menu to the menu bar
        menubar.addAction(filemenu.menuAction())
        menubar.setGeometry(0, 0, 800, 22)
        filemenu.setFixedSize(240, 100)

        # Save all spectra "button"
        self.save_all_button = QtGui.QAction("Save All Spectra As...")
        self.save_all_button.triggered.connect(self.saveAllSpectra)
        self.save_all_button.setDisabled(True)
        self.save_all_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_all_button)
        filemenu.addSeparator()

        # Save selected spectra "button"
        self.save_disp_button = QtGui.QAction("Save Selected Spectra As...")
        self.save_disp_button.triggered.connect(self.saveDispSpectra)
        self.save_disp_button.setDisabled(True)
        self.save_disp_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_disp_button)
        filemenu.addSeparator()

        # Save average spectrum "button"
        self.save_avg_button = QtGui.QAction("Save Average Spectrum As...")
        self.save_avg_button.triggered.connect(self.saveAvgSpectrum)
        self.save_avg_button.setDisabled(True)
        self.save_avg_button.setIcon(QtGui.QIcon("icons/save-icon.png"))
        filemenu.addAction(self.save_avg_button)

        # Energy map selection box
        self.emap_combo = QtWidgets.QComboBox()
        self.emap_combo.setFixedSize(140, 30)
        if self.emap is not None:
            self.emap_combo.addItem(self.emap.name)

        # Load Energy Map button
        self.emap_load_button = QtWidgets.QPushButton("Load Energy Map...")
        self.emap_load_button.setFixedSize(140, 30)
        self.emap_load_button.clicked.connect(self.loadEmap)

        # Gradient type (or rainbow) dropdown box (combo box)
        self.colour_box = QtWidgets.QComboBox()
        self.colour_box.setFixedSize(140, 30)
        self.colour_box.addItem("Red Gradient")
        self.colour_box.addItem("Red Gradient (inverted)")
        self.colour_box.addItem("Green Gradient")
        self.colour_box.addItem("Green Gradient (inverted)")
        self.colour_box.addItem("Blue Gradient")
        self.colour_box.addItem("Blue Gradient (inverted)")
        self.colour_box.addItem("Black-White Gradient")
        self.colour_box.addItem("Black-White (inverted)")
        self.colour_box.addItem("Rainbow")
        self.colour_box.addItem("Custom Gradient")
        self.custom_colour_one = QtGui.QColor(0, 0, 0)
        self.custom_colour_two = QtGui.QColor(255, 255, 255)
        self.colour_box.setCurrentIndex(3)

        # Set custom colours button
        self.custom_col_button = QtWidgets.QPushButton("Set Custom Colours")
        self.custom_col_button.clicked.connect(self.setCustomColours)
        self.custom_col_button.setFixedSize(140, 30)

        # Type of "stacking" of spectra (dropdown box, aka combo box)
        self.stack_type_box = QtWidgets.QComboBox()
        self.stack_type_box.setFixedSize(140, 30)
        self.stack_type_box.addItem("Stacked (no spacing)")
        self.stack_type_box.addItem("Spaced")
        self.stack_type_box.addItem("Average of Spectra")
        self.stack_type_box.setCurrentIndex(1)

        # Refresh button
        self.refresh_button = QtWidgets.QPushButton("Refresh")
        self.refresh_button.setFixedSize(140, 30)
        self.refresh_button.setDisabled(True)
        self.refresh_button.clicked.connect(self.refreshSpectra)

        # connects everything to the XES window
        widget = QtWidgets.QWidget()
        self.mlayout = QtWidgets.QGridLayout(widget)
        self.mlayout.addWidget(self.sc, 2, 2, 1, 2, AlignFlag.AlignCenter)
        self.mlayout.addWidget(load_xes_button, 0, 0, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.colour_box, 0, 1, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.custom_col_button, 0, 2, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.emap_combo, 0, 3, AlignFlag.AlignRight)
        self.mlayout.addWidget(self.stack_type_box, 1, 1, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.refresh_button, 1, 2, AlignFlag.AlignLeft)
        self.mlayout.addWidget(self.emap_load_button, 1, 3, AlignFlag.AlignRight)
        self.mlayout.setColumnMinimumWidth(2, 500)

        self.setCentralWidget(widget)

        self.show()

    # handles close event so a confirmation window can appear
    def closeEvent(self, event):
        # The no_close_dialog exists so the window can be closed by a MainWindow with no issue
        if not self.no_close_dialog:
            if hasattr(self.parent, "confirm_on_close"):
                if self.parent.confirm_on_close:
                    confirm = exitDialog(self)
                else:
                    confirm = True
            else:
                confirm = exitDialog(self)
            if confirm:
                if hasattr(self.parent, "childWindow"):
                    self.parent.childWindow = None
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    # loads XES data (currently only able to load from TIF files)
    def loadXES(self):
        if len(self.emaps):
            self.emap = self.emaps[self.emap_combo.currentIndex()]
        elif self.emap is None:
            try:
                self.emap = self.parent.emap
            except Exception:
                self.error = ErrorWindow("XESemap")
                return

        dtype = self.loadType()
        if dtype == "tif":
            self.filenames = LoadTifSpectraData.fileDialog(self)
        elif dtype == "h5py":
            self.filenames = LoadH5Data.fileDialog(self)
        else:
            raise TypeError(f"unknown dtype {dtype}")

        if not self.filenames:
            return

        emap = self.emap
        # gets all XES spectra.
        LoadWindow = LoadingBarWindow("Loading XES data...", len(self.filenames))
        scanset = []
        data = calcDataForSpectra(emap)
        hnames = {}
        for i in self.filenames:
            if LoadWindow.wasCanceled():
                break
            spectra = calcSpectra(i, emap, data, dtype)
            if type(spectra) is list:
                hnames[i] = len(spectra)
                scanset += spectra
            else:
                scanset.append(spectra)
            LoadWindow.add()
            QtWidgets.QApplication.processEvents()
        LoadWindow.deleteLater()

        if LoadWindow.wasCanceled():
            return

        self.checks = QtWidgets.QScrollArea()
        self.checks.setMinimumWidth(280)
        self.check_widgets = QtWidgets.QWidget()
        self.checks_grid = QtWidgets.QGridLayout(self.check_widgets)

        # scanset = calcXESSpectra(self.filenames, emap)
        scanlen = len(scanset)
        colour_index = self.colour_box.currentIndex()
        if colour_index == 9:
            cols = (self.custom_colour_one.getRgb(), self.custom_colour_two.getRgb())
            colours = colourGen(scanlen, None, cols, True)
        else:
            colours = colourGen(scanlen, colour_index)
        if dtype == "h5py":
            names = []
            for i in hnames:
                n = hnames[i]
                names += [i + f"-{j}" for j in range(n)]

            self.spectra = [
                Spectrum(self, scanset[i], colours[i], i + 3, names[i])
                for i, _ in enumerate(scanset)
            ]
        else:
            self.spectra = [
                Spectrum(
                    self,
                    scanset[i],
                    colours[i],
                    i + 3,
                )
                for i, _ in enumerate(scanset)
            ]

        self.disp_spectra = self.spectra.copy()

        self.refresh_button.setDisabled(False)

        all_button = QtWidgets.QPushButton()
        all_button.clicked.connect(self.allSpectra)
        all_button.setText("All")
        none_button = QtWidgets.QPushButton()
        none_button.clicked.connect(self.noSpectra)
        none_button.setText("None")
        invert_button = QtWidgets.QPushButton()
        invert_button.clicked.connect(self.invertSpectra)
        invert_button.setText("Invert")
        self.checks_grid.addWidget(all_button, 0, 0, AlignFlag.AlignLeft)
        self.checks_grid.addWidget(none_button, 0, 1, AlignFlag.AlignLeft)
        self.checks_grid.addWidget(invert_button, 0, 2, AlignFlag.AlignLeft)
        self.checks.setWidget(self.check_widgets)
        self.mlayout.addWidget(
            self.checks, 2, 0, 1, 2, AlignFlag.AlignHCenter | AlignFlag.AlignTop
        )

        self.save_all_button.setDisabled(False)
        self.save_disp_button.setDisabled(False)
        self.save_avg_button.setDisabled(False)

        self.refreshSpectra()

    # sets colours for custom gradient
    def setCustomColours(self):
        self.ColourSelects = ColourSelect(
            self, self.custom_colour_one, self.custom_colour_two
        )

    # runs stack and graph in one function (saves space)
    def refreshSpectra(self):
        self.stackSpectra()
        self.graphSpectra()

    # selects all spectra
    def allSpectra(self):
        for i in self.spectra:
            i.restack_now = False
            i.box.setChecked(True)
            i.restack_now = True
        self.refreshSpectra()

    # deselects all spectra
    def noSpectra(self):
        for i in self.spectra:
            i.restack_now = False
            i.box.setChecked(False)
            i.restack_now = True
        self.refreshSpectra()

    # inverts spectra selection
    def invertSpectra(self):
        for i in self.spectra:
            i.restack_now = False
            if i in self.disp_spectra:
                i.box.setChecked(False)
            else:
                i.box.setChecked(True)
            i.restack_now = True
        self.refreshSpectra()

    # calculates the average spectra
    def setAverageSpectra(self):
        energies = []
        intensities = []
        for i, _ in enumerate(self.disp_spectra[0].energies):
            energies.append(
                sum(x.energies[i] for x in self.disp_spectra) / len(self.disp_spectra)
            )
        for i, _ in enumerate(self.disp_spectra[0].intensities):
            intensities.append(
                sum(x.intensities[i] for x in self.disp_spectra)
                / len(self.disp_spectra)
            )
        return (energies, intensities)

    # does the math to find how the spectra should be drawn
    def stackSpectra(self):
        stack_type = self.stack_type_box.currentIndex()
        if stack_type == 2:
            if not len(self.disp_spectra):
                self.average_spectra = None
            else:
                self.average_spectra = self.setAverageSpectra()

        else:
            self.average_spectra = None
            if stack_type:
                amt = 100
            else:
                amt = 0
            for index, s in enumerate(self.disp_spectra):
                s.current = s.base.copy()
                for j, _ in enumerate(s.current):
                    inc = amt * index
                    s.increaseIntensity(inc, j)

    # draws the spectra to the figure
    def graphSpectra(self):
        if self.average_spectra is not None:
            s = self.average_spectra
            self.sc.plotItem.clear()
            self.sc.plotItem.plot(s[0], s[1], pen=pg.mkPen(color="k", width=2))

        else:
            coltype = self.colour_box.currentIndex()
            if coltype == 9:
                cols = (
                    self.custom_colour_one.getRgb(),
                    self.custom_colour_two.getRgb(),
                )
                colours = colourGen(len(self.spectra), None, cols, True)
                for i, s in enumerate(self.spectra):
                    s.colour = colours[i]
                    s.editBoxText()
            else:
                colours = colourGen(len(self.spectra), coltype)
                for i, s in enumerate(self.spectra):
                    s.colour = colours[i]
                    s.editBoxText()
            self.sc.plotItem.clear()
            for i in self.disp_spectra:
                self.sc.plotItem.plot(
                    i.energies, i.current, pen=pg.mkPen(color=i.colour, width=2)
                )

    # disables given spectrum
    def removeSpectrum(self, spectrum):
        self.disp_spectra.remove(spectrum)
        if spectrum.restack_now:
            self.refreshSpectra()

    # enables given spectrum
    def addSpectrum(self, spectrum):
        num = self.spectra.index(spectrum)
        count = 0
        for i in range(num):
            if self.spectra[i] in self.disp_spectra:
                count += 1
        self.disp_spectra.insert(count, spectrum)
        if spectrum.restack_now:
            self.refreshSpectra()

    # saves given spectra (or all if none are given)
    def saveSpectra(self, spectra=None):
        if spectra is None:
            spectra = self.spectra

        dialog = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save All Spectra",
            filter=("Excel Spreadsheet (*.xlsx)\nSimple Text Layout (*.csv)"),
        )
        if dialog[1] == "Excel Spreadsheet (*.xlsx)":
            wb = ExWorkbook()
            ws = wb.active

            lines = [[] for _, _ in enumerate(spectra[0].energies)]
            lines.insert(0, [])
            lines.insert(1, [])
            for spect in spectra:
                lines[0].append(spect.name[: spect.name.rfind(".")])
                lines[0].append("")
                lines[0].append("")
                lines[1].append("Emission Energy (eV)")
                lines[1].append("Counts")
                lines[1].append("")
                texts = [
                    [str(spect.energies[j]), str(spect.intensities[j]), ""]
                    for j, _ in enumerate(spect.energies)
                ]
                for k, item in enumerate(texts):
                    for l in range(3):
                        lines[k + 2].append(item[l])
            for i, line in enumerate(lines, 1):
                for j, item in enumerate(line, 1):
                    try:
                        n = float(item)
                    except Exception:
                        n = item
                    ws.cell(i, j).value = n
            for i in range(int(len(lines[0]) / 3)):
                ws.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=i * 3 + 1,
                    end_column=i * 3 + 2,
                )
                ws.column_dimensions[getColumnLetter(i * 3 + 1)].width = 20
            wb.save(dialog[0])
            wb.close()

        elif dialog[1] == "Simple Text Layout (*.csv)":
            direct = open(dialog[0], "+w")
            direct.seek(0)
            direct.truncate()
            lines = ["" for _, _ in enumerate(spectra[0].energies)]
            lines.insert(0, "")
            lines.insert(1, "")
            for spect in spectra:
                lines[0] += spect.name + ",,,"
                lines[1] += "Emission Energy (eV),Counts,,"
                texts = [
                    str(spect.energies[j]) + "," + str(spect.intensities[j]) + ",,"
                    for j, _ in enumerate(spect.energies)
                ]
                for k, string in enumerate(texts):
                    lines[k + 2] += string
            text = ""
            for line in lines:
                text += line + "\n"
            direct.write(text)
            direct.close()

    # runs saveSpectra with all selected
    def saveAllSpectra(self):
        self.saveSpectra()

    # saves enables spectra only
    def saveDispSpectra(self):
        if not len(self.disp_spectra):
            self.error = ErrorWindow("nodispSpec")
        self.saveSpectra(self.disp_spectra)

    # saves the average spectrum (recalculates before saving)
    def saveAvgSpectrum(self):
        try:
            self.average_spectra = self.setAverageSpectra()
        except IndexError:
            self.error = ErrorWindow("avgNoSelected")
        self.saveSpectra(self.average_spectra)


# creates an XES window when file is run
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setWindowIcon(QtGui.QIcon("icons/spc-logo-nobg.png"))
    W = XESWindow(None)
    # sys.exit(app.exec())
    app.exec()
